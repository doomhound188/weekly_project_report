"""
SharePoint project roadmap parser and comparison logic.
"""

import os
from typing import Optional
from openpyxl import load_workbook


class SharePointProjects:
    """Parses SharePoint project roadmap and compares with timesheet entries."""

    def __init__(self, excel_path: str):
        """
        Initialize with path to the downloaded Excel file.
        
        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.projects = []

    def parse_roadmap(self) -> list[dict]:
        """
        Parse the project roadmap Excel file.
        
        Returns:
            List of project dictionaries with name, status, dates, etc.
        """
        if not os.path.exists(self.excel_path):
            print(f"⚠️  Excel file not found: {self.excel_path}")
            return []

        try:
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            
            # Try to find the main sheet (adjust sheet name as needed)
            sheet_names = wb.sheetnames
            sheet = None
            
            # Look for common sheet names
            for name in ["Projects", "Roadmap", "Active", "2025", "Sheet1"]:
                if name in sheet_names:
                    sheet = wb[name]
                    break
            
            if sheet is None:
                sheet = wb.active

            # Parse the header row to find column indices
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[str(cell.value).lower().strip()] = col_idx

            # Common column name variations
            name_cols = ["project", "project name", "name", "title"]
            status_cols = ["status", "project status", "state"]
            owner_cols = ["owner", "assigned to", "assigned", "resource", "technician"]
            start_cols = ["start", "start date", "begin"]
            end_cols = ["end", "end date", "due", "due date", "target"]

            def find_col(options):
                for opt in options:
                    if opt in headers:
                        return headers[opt]
                return None

            name_col = find_col(name_cols)
            status_col = find_col(status_cols)
            owner_col = find_col(owner_cols)
            start_col = find_col(start_cols)
            end_col = find_col(end_cols)

            if not name_col:
                print("⚠️  Could not find project name column in Excel")
                return []

            # Parse project rows
            projects = []
            for row in sheet.iter_rows(min_row=2):
                name_cell = row[name_col - 1] if name_col else None
                
                if not name_cell or not name_cell.value:
                    continue

                project = {
                    "name": str(name_cell.value).strip(),
                    "status": None,
                    "owner": None,
                    "start_date": None,
                    "end_date": None
                }

                if status_col and len(row) >= status_col:
                    val = row[status_col - 1].value
                    project["status"] = str(val).strip() if val else None

                if owner_col and len(row) >= owner_col:
                    val = row[owner_col - 1].value
                    project["owner"] = str(val).strip() if val else None

                if start_col and len(row) >= start_col:
                    project["start_date"] = row[start_col - 1].value

                if end_col and len(row) >= end_col:
                    project["end_date"] = row[end_col - 1].value

                projects.append(project)

            wb.close()
            self.projects = projects
            print(f"✅ Parsed {len(projects)} projects from roadmap")
            return projects

        except Exception as e:
            print(f"❌ Error parsing Excel file: {e}")
            return []

    def get_active_projects(self, owner_filter: Optional[str] = None) -> list[dict]:
        """
        Get active/open projects, optionally filtered by owner.
        
        Args:
            owner_filter: Optional owner name to filter by (case-insensitive partial match)
            
        Returns:
            List of active project dictionaries
        """
        if not self.projects:
            self.parse_roadmap()

        active_statuses = ["active", "in progress", "open", "current", "ongoing", None]
        
        active_projects = []
        for project in self.projects:
            status = project.get("status", "").lower() if project.get("status") else ""
            
            # Check if status indicates active
            is_active = any(s in status for s in ["active", "progress", "open", "current"]) or not status
            
            # Skip completed/closed projects
            if any(s in status for s in ["complete", "closed", "done", "cancelled", "canceled"]):
                continue

            # Filter by owner if specified
            if owner_filter:
                owner = project.get("owner", "") or ""
                if owner_filter.lower() not in owner.lower():
                    continue

            if is_active:
                active_projects.append(project)

        return active_projects

    def compare_with_timesheet(
        self, 
        timesheet_projects: list[str], 
        owner_filter: Optional[str] = None
    ) -> dict:
        """
        Compare roadmap projects with timesheet entries.
        
        Args:
            timesheet_projects: List of project names from timesheet
            owner_filter: Optional owner name to filter roadmap by
            
        Returns:
            Dictionary with comparison results:
            - in_both: Projects in both roadmap and timesheet
            - roadmap_only: Projects in roadmap but not timesheet (missing from report)
            - timesheet_only: Projects in timesheet but not roadmap
        """
        roadmap_projects = self.get_active_projects(owner_filter)
        roadmap_names = {p["name"].lower(): p for p in roadmap_projects}
        timesheet_lower = {name.lower(): name for name in timesheet_projects}

        in_both = []
        roadmap_only = []
        timesheet_only = []

        # Check roadmap projects
        for name_lower, project in roadmap_names.items():
            # Fuzzy match: check if roadmap project name is contained in any timesheet project
            matched = False
            for ts_lower, ts_name in timesheet_lower.items():
                if name_lower in ts_lower or ts_lower in name_lower:
                    in_both.append({
                        "roadmap": project["name"],
                        "timesheet": ts_name
                    })
                    matched = True
                    break
            
            if not matched:
                roadmap_only.append(project)

        # Check timesheet projects not in roadmap
        for ts_lower, ts_name in timesheet_lower.items():
            matched = any(
                name_lower in ts_lower or ts_lower in name_lower 
                for name_lower in roadmap_names.keys()
            )
            if not matched:
                timesheet_only.append(ts_name)

        return {
            "in_both": in_both,
            "roadmap_only": roadmap_only,
            "timesheet_only": timesheet_only
        }

    def format_comparison_summary(self, comparison: dict) -> str:
        """
        Format the comparison results as a summary string.
        
        Args:
            comparison: Dictionary from compare_with_timesheet
            
        Returns:
            Formatted summary string
        """
        lines = []
        
        if comparison["roadmap_only"]:
            lines.append("\n⚠️  Projects in Roadmap but MISSING from this week's report:")
            for project in comparison["roadmap_only"]:
                end_date = project.get("end_date", "N/A")
                if hasattr(end_date, "strftime"):
                    end_date = end_date.strftime("%B %d")
                lines.append(f"   - {project['name']} (Due: {end_date})")

        if comparison["timesheet_only"]:
            lines.append("\n📝 Projects with time entries but NOT in Roadmap:")
            for name in comparison["timesheet_only"]:
                lines.append(f"   - {name}")

        if not lines:
            lines.append("\n✅ All roadmap projects are accounted for in this week's report.")

        return "\n".join(lines)
